import os
from io import BytesIO
from datetime import datetime, date
from decimal import Decimal

# 🧠 Python Standard + Django
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.text import slugify
from django.conf import settings
from django.contrib import messages
from django.db.models import Count, Q, F, ProtectedError
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView, ListView, FormView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import LogoutView as DjangoLogoutView
from django.contrib.auth.views import LoginView as DjangoLoginView
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.models import User

# 📦 Third-party
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.styles import ParagraphStyle
from svglib.svglib import svg2rlg
import arabic_reshaper
from bidi.algorithm import get_display

# 🧩 Local
from .forms import (
    ProfileUpdateForm, PasswordChangeForm, ContractForm, ContractItemFormSet, ZoneFormSet, WarehouseForm, TaskForm,
    DeviceCategoryForm, DeviceForm, DevicePropertyFormSet,
    MaintenanceCardForm, CoordinationRequestForm
)

from .models import (
    Contract, DeviceCategory, ContractItem, Warehouse, Zone,
    Task, Device, DeviceProperty, MaintenanceCard, CoordinationRequest
)


def custom_403(request, exception):
    return render(request, '403.html', {}, status=403)

def custom_404(request, exception):
    return render(request, '404.html', {}, status=404)

def custom_500(request):
    return render(request, '500.html', status=500)

# Export Excels & PDFs
class ExportMixin:
    def get(self, request, *args, **kwargs):
        if "export" in request.GET:
            export_format = request.GET.get("format", "excel")
            queryset = self.get_queryset()

            if not queryset.exists():
                return HttpResponse("لا توجد بيانات للتصدير", content_type="text/plain")

            if export_format == "pdf":
                return self.export_to_pdf(queryset)
            else:
                return self.export_to_excel(queryset)

        return super().get(request, *args, **kwargs)

    def prepare_export_response(self, content, file_type, model_name):
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        safe_model_name = slugify(model_name) or "export"
        filename = f"export_{safe_model_name}_{timestamp}.{file_type}"

        content_types = {
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "pdf": "application/pdf",
        }

        response = HttpResponse(content, content_type=content_types.get(file_type, "application/octet-stream"))
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def export_to_excel(self, queryset):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"{queryset.model._meta.verbose_name_plural}"

        fields = [field.name for field in queryset.model._meta.fields[1:]]
        headers = [field.verbose_name for field in queryset.model._meta.fields[1:]]

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            sheet.column_dimensions[get_column_letter(col_num)].width = max(15, len(header) + 2)

        for obj in queryset:
            row = []
            for field in fields:
                value = getattr(obj, field, "")
                if hasattr(value, "get_FOO_display"):
                    value = value.get_FOO_display()
                elif hasattr(value, "__str__"):
                    value = str(value)
                row.append(value)
            sheet.append(row)

        # Apply border to all cells
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                                   min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = border

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return self.prepare_export_response(output.read(), "xlsx", queryset.model._meta.model_name)

    def export_to_pdf(self, queryset):
        model_name = queryset.model._meta.model_name
        buffer = BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=10,
            leftMargin=10,
            topMargin=4,
            bottomMargin=20
        )

        font_path = os.path.join(settings.STATIC_ROOT, "contracts/fonts/Janna LT Bold/Janna LT Bold.ttf")
        pdfmetrics.registerFont(TTFont("Janna", font_path))

        elements = []

        # ✅ إضافة صورة SVG
        svg_path = os.path.join(settings.STATIC_ROOT, "contracts/img/logo-1.svg")

        drawing = svg2rlg(svg_path)
        drawing.scale(1, 1)  # التحكم بالحجم
        drawing.hAlign = "LEFT"

        # إعداد العنوان
        title_text = get_display(arabic_reshaper.reshape(f"تقرير {queryset.model._meta.verbose_name_plural}"))
        title_style = ParagraphStyle(name="Title", fontName="Janna", fontSize=20, alignment=1, spaceAfter=0)
        title_paragraph = Paragraph(title_text, title_style)

        # إنشاء جدول مكون من عمودين: [الصورة, العنوان]
        title_table = Table(
            data=[[drawing, title_paragraph]],
            colWidths=[80, 650],  # يمكنك تعديل العرض حسب حجم الصورة والعنوان
            hAlign='RIGHT'
        )
        title_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, 0), "RIGHT"),  # الصورة
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),  # العنوان
        ]))

        elements.append(title_table)
        elements.append(Spacer(1, 25))

        # ✅ جدول البيانات
        fields = [[field.name, field.verbose_name] for field in queryset.model._meta.fields[-1:0:-1]]
        headers = [get_display(arabic_reshaper.reshape(field[1])) for field in fields]
        table_data = [headers]
        max_col_lengths = [len(h) for h in headers]

        for obj in queryset:
            row = []
            for i, field in enumerate(fields):
                value = getattr(obj, field[0], "")
                if hasattr(obj, f"get_{field[0]}_display"):
                    value = getattr(obj, f"get_{field[0]}_display")()
                elif isinstance(value, (datetime, date)):
                    value = value.strftime("%Y-%m-%d")
                elif isinstance(value, bool):
                    value = "نعم" if value else "لا"
                else:
                    value = str(value)

                if any("\u0600" <= c <= "\u06FF" for c in value):
                    value = get_display(arabic_reshaper.reshape(value))

                max_col_lengths[i] = max(max_col_lengths[i], len(value))
                row.append(value)
            table_data.append(row)

        # ✅ تحديد عرض الأعمدة تلقائيًا
        total_width = 780
        min_width = 55
        max_width = 220 
        col_widths = []
        sum_lengths = sum(max_col_lengths) or 1 
        col_widths = [
          max(min_width, min((length / sum_lengths) * total_width, max_width)) 
          for length in max_col_lengths
        ]

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, -1), "Janna"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 16))

        doc.build(elements)
        return self.prepare_export_response(buffer.getvalue(), "pdf", model_name)

class AuthViewMixin(LoginRequiredMixin):
    login_url = 'login'  # تأكد من توفر هذا الـ URL

class LoginView(DjangoLoginView):
    template_name = 'registration/login.html'

    def get_success_url(self):
        return reverse_lazy('dashboard')

class LogoutView(AuthViewMixin, DjangoLogoutView):
    next_page = reverse_lazy('login')

    def get(self, request, *args, **kwargs):
        return self.post(request, *args, **kwargs)

class ProfileView(AuthViewMixin, TemplateView):
    template_name = 'registration/profile.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['username_form'] = ProfileUpdateForm(instance=self.request.user)
        context['password_form'] = PasswordChangeForm()
        return context

    def post(self, request, *args, **kwargs):
        if 'update_username' in request.POST:
            form = ProfileUpdateForm(request.POST, instance=request.user)
            if form.is_valid():
                form.save()
                messages.success(request, "Username updated successfully.")
            else:
                messages.error(request, "Failed to update username.")
        elif 'change_password' in request.POST:
            form = PasswordChangeForm(request.POST)
            if form.is_valid():
                user = request.user
                if not user.check_password(form.cleaned_data['current_password']):
                    messages.error(request, "Current password is incorrect.")
                else:
                    user.set_password(form.cleaned_data['new_password'])
                    user.save()
                    update_session_auth_hash(request, user)  # Keep user logged in
                    messages.success(request, "Password changed successfully.")
            else:
                messages.error(request, "Failed to change password.")
        return redirect('profile')


class DashboardView(AuthViewMixin, TemplateView):
    template_name = 'contracts/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # عقود
        contracts = Contract.objects.all()
        total_contracts = contracts.count()
        latest_contracts = contracts.order_by('-start_date')[:5]

        # مناطق
        total_zones = Zone.objects.count()

        # أجهزة
        devices = Device.objects.all()
        total_devices = devices.count()
        installed_devices = devices.filter(status='installed').count()
        available_devices = devices.filter(status='available').count()
        damaged_devices = devices.filter(status='damaged').count()
        devices_in_warehouse = devices.filter(current_location='warehouse').count()

        # مهام
        tasks = Task.objects.all()
        task_summary = {
            'total': tasks.count(),
            'completed': tasks.filter(status='completed').count(),
            'not_started': tasks.filter(status='not_started').count(),
            'ongoing': tasks.filter(status='ongoing').count(),
            'delayed': tasks.filter(status='delayed').count(),
        }

        # صيانة
        maintenance = MaintenanceCard.objects.all()
        maintenance_summary = {
            'total': maintenance.count(),
            'repaired': maintenance.exclude(repair_date__isnull=True).count(),
            'pending': maintenance.filter(repair_date__isnull=True).count(),
        }

        # تنسيق
        coordination = CoordinationRequest.objects.all()
        coordination_summary = {
            'total': coordination.count(),
            'last_date': coordination.order_by("-request_date").first().request_date if coordination.exists() else "—"
        }
        latest_tasks = Task.objects.filter(status__in=["not_started", "ongoing"]).order_by('-deadline')[:20]

        context.update({
            "total_contracts": total_contracts,
            "latest_contracts": latest_contracts,
            "total_zones": total_zones,
            "total_devices": total_devices,
            "installed_devices": installed_devices,
            "available_devices": available_devices,
            "damaged_devices": damaged_devices,
            "devices_in_warehouse": devices_in_warehouse,
            "task_summary": task_summary,
            "maintenance_summary": maintenance_summary,
            "coordination_summary": coordination_summary,
            'latest_tasks': latest_tasks,
        })
        return context

class ContractListView(AuthViewMixin, ListView):
    model = Contract
    template_name = 'contracts/contracts/list.html'
    context_object_name = 'contracts'

class ContractDeleteView(AuthViewMixin, DeleteView):
    model = Contract
    template_name = 'contracts/contracts/delete.html'
    context_object_name = 'contracts'
    success_url = reverse_lazy('contract_list')
    
    def delete(self, request, *args, **kwargs):
        contract = self.get_object()
        messages.success(request, f"Contract '{contract.name}' was deleted successfully.")
        return super().delete(request, *args, **kwargs)

class ContractDetailView(AuthViewMixin, DetailView):
    model = Contract
    template_name = 'contracts/contracts/detail.html'
    context_object_name = 'contract'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contract = self.object

        context["items"] = contract.items.all()
        context["zones"] = contract.zones.all()
        context["warehouse"] = getattr(contract, 'warehouse', None)
        context["devices"] = Device.objects.filter(zone__contract=contract)
        context["maintenance_cards"] = MaintenanceCard.objects.filter(device__zone__contract=contract)
        context["coordination_requests"] = CoordinationRequest.objects.filter(zone__contract=contract)
        context["tasks"] = Task.objects.filter(zone__contract=contract)

        category_data = []
        for item in context["items"]:
            device_count = Device.objects.filter(device_category=item.category, zone__contract=contract).count()
            percentage = (device_count / item.quantity * 100) if item.quantity else 0
            category_data.append({
                "category_label": item.category.name,
                "required": item.quantity,
                "available": device_count,
                "percentage": round(percentage, 1)
            })
        context["category_stats"] = category_data

        zone_stats = []
        for zone in context["zones"]:
            devices = zone.devices.all()
            zone_stats.append({
                'zone': zone,
                'installed': devices.filter(status='installed').count(),
                'available': devices.filter(status='available').count(),
                'damaged': devices.filter(status='damaged').count(),
                'total': devices.count(),
            })
        context["zone_stats"] = zone_stats

        tasks = context["tasks"]
        context["task_summary"] = {
            'total': tasks.count(),
            'completed': tasks.filter(status='completed').count(),
            'ongoing': tasks.filter(status='ongoing').count(),
            'not_started': tasks.filter(status='not_started').count(),
            'delayed': tasks.filter(status='delayed').count(),
        }

        maintenance = context["maintenance_cards"]
        context["maintenance_summary"] = {
            'total': maintenance.count(),
            'repaired': maintenance.exclude(repair_date__isnull=True).count(),
            'pending': maintenance.filter(repair_date__isnull=True).count(),
        }

        coordination = context["coordination_requests"]
        context["coordination_summary"] = {
            'total': coordination.count(),
            'last_date': coordination.order_by("-request_date").first().request_date if coordination.exists() else "—"
        }

        return context

class ContractFormView(AuthViewMixin, CreateView, UpdateView):
    model = Contract
    form_class = ContractForm
    template_name = 'contracts/contracts/form.html'
    success_url = reverse_lazy('contract_list')

    def get_object(self, queryset=None):
        pk = self.kwargs.get('pk')
        if pk:
            return get_object_or_404(Contract, pk=pk)
        return None

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contract = self.get_object()

        if self.request.POST:
            context['item_formset'] = ContractItemFormSet(self.request.POST, instance=contract, prefix='items')
            context['zone_formset'] = ZoneFormSet(self.request.POST, instance=contract, prefix='zones')
            context['warehouse_form'] = WarehouseForm(self.request.POST, instance=getattr(contract, 'warehouse', None), prefix='warehouse')
        else:
            context['item_formset'] = ContractItemFormSet(instance=contract, prefix='items')
            context['zone_formset'] = ZoneFormSet(instance=contract, prefix='zones')
            context['warehouse_form'] = WarehouseForm(instance=getattr(contract, 'warehouse', None), prefix='warehouse')

        return context

    def form_valid(self, form):
        context = self.get_context_data()
        item_formset = context['item_formset']
        zone_formset = context['zone_formset']
        warehouse_form = context['warehouse_form']

        forms_valid = all([
            form.is_valid(),
            item_formset.is_valid(),
            zone_formset.is_valid(),
            warehouse_form.is_valid(),
        ])

        if forms_valid:
            self.object = form.save()
            item_formset.instance = self.object
            item_formset.save()

            zone_formset.instance = self.object
            zone_formset.save()

            warehouse = warehouse_form.save(commit=False)
            warehouse.contract = self.object
            warehouse.save()

            messages.success(self.request, "Contract, items, zones, and warehouse saved successfully.")
            return HttpResponseRedirect(self.get_success_url())

        if not item_formset.is_valid():
            messages.error(self.request, "There are errors in the contract items:")
            for form in item_formset:
                for field, errors in form.errors.items():
                    messages.error(self.request, f"• {field}: {errors.as_text()}")

        if not zone_formset.is_valid():
            messages.error(self.request, "There are errors in the contract zones:")
            for form in zone_formset:
                for field, errors in form.errors.items():
                    messages.error(self.request, f"• {field}: {errors.as_text()}")

        if not warehouse_form.is_valid():
            messages.error(self.request, "There are errors in the warehouse form:")
            for field, errors in warehouse_form.errors.items():
                messages.error(self.request, f"• {field}: {errors.as_text()}")

        return self.form_invalid(form)

    def form_invalid(self, form):
        messages.error(self.request, "An error occurred while saving the contract.")
        return super().form_invalid(form)

@csrf_exempt
def manage_device_categories(request):
    categories = DeviceCategory.objects.all().order_by('id')

    if request.method == "POST":
        if 'save' in request.POST:
            category_id = request.POST.get("category_id")
            if category_id:
                category = get_object_or_404(DeviceCategory, id=category_id)
                form = DeviceCategoryForm(request.POST, instance=category)
                if form.is_valid():
                    form.save()
                    messages.success(request, "Category updated successfully.")
                    return redirect('manage_device_categories')
            else:
                form = DeviceCategoryForm(request.POST)
                if form.is_valid():
                    form.save()
                    messages.success(request, "Category added successfully.")
                    return redirect('manage_device_categories')

        elif 'delete' in request.POST:
            category_id = request.POST.get("category_id")
            category = get_object_or_404(DeviceCategory, id=category_id)
            try:
                category.delete()
                messages.success(request, "Category deleted successfully.")
            except ProtectedError:
                messages.error(request, f"Cannot delete category '{category.name}' because it is used by other records.")
            return redirect('manage_device_categories')

    form = DeviceCategoryForm()
    return render(request, 'contracts/device_categories/manage.html', {
        'categories': categories,
        'form': form,
    })

class WarehouseListView(AuthViewMixin, ListView):
    model = Warehouse
    template_name = 'contracts/warehouses/list.html'
    context_object_name = 'warehouses'

class WarehouseDetailView(AuthViewMixin, DetailView):
    model = Warehouse
    template_name = 'contracts/warehouses/detail.html'
    context_object_name = 'warehouse'

    def get_queryset(self):
        return Warehouse.objects.select_related("contract").prefetch_related("devices")

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()

        if request.GET.get("export") == "true":
            format_ = request.GET.get("format")
            queryset = self.get_filtered_devices()
            if format_ == "excel":
                return self.export_to_excel(queryset)
            elif format_ == "pdf":
                return self.export_to_pdf(queryset)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        warehouse = self.object

        # فلترة
        devices = self.get_filtered_devices()
        context['devices'] = devices

        context['filter_status'] = self.request.GET.get('status')
        context['filter_zone'] = self.request.GET.get('zone')
        context['search_query'] = self.request.GET.get('q')
        context['zones'] = Zone.objects.filter(contract=warehouse.contract)

        context['total_devices'] = devices.count()
        context['category_counts'] = devices.values('device_category__name').annotate(total=Count('serial_number'))
        return context

    def export_to_pdf(self, queryset):
        warehouse = self.object
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)

        font_path = os.path.join(settings.BASE_DIR, "static", "contracts", "fonts", "Janna LT Bold", "Janna LT Bold.ttf")
        pdfmetrics.registerFont(TTFont("Janna", font_path))

        content = []

        # ✅ إضافة صورة SVG
        svg_path = os.path.join(settings.STATIC_ROOT, "contracts/img/logo-1.svg")

        drawing = svg2rlg(svg_path)
        drawing.scale(1, 1)  # التحكم بالحجم
        drawing.hAlign = "LEFT"

        # إعداد العنوان
        contract_info = f"تقرير أجهزة العقد  : {warehouse.contract.name} - {warehouse.contract.contract_number}"
        title_text = get_display(arabic_reshaper.reshape(contract_info))
        title_style = ParagraphStyle(name="Title", fontName="Janna", fontSize=20, alignment=1, spaceAfter=0)
        title_paragraph = Paragraph(title_text, title_style)

        # إنشاء جدول مكون من عمودين: [الصورة, العنوان]
        title_table = Table(
            data=[[drawing, title_paragraph]],
            colWidths=[80, 650],  # يمكنك تعديل العرض حسب حجم الصورة والعنوان
            hAlign='RIGHT'
        )
        title_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, 0), "RIGHT"),  # الصورة
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),  # العنوان
        ]))

        content.append(title_table)
        content.append(Spacer(1, 25))

        
        title_text = f""
        if any("\u0600" <= c <= "\u06FF" for c in title_text):
            title_text = get_display(arabic_reshaper.reshape(title_text))

        title = Paragraph(title_text, ParagraphStyle("title", fontName="Janna", fontSize=14, alignment=1))
        content.append(title)
        content.append(Spacer(1, 10))

        # إنشاء الفقرات
        stats = [
            f"عدد الأجهزة في المخزن: {warehouse.count_in_warehouse}",
            f"المخزن: {warehouse.name} - {warehouse.location}",
            f"إجمالي المناطق: {warehouse.count_zones}",
            f"إجمالي الأجهزة: {warehouse.count_devices}",
            f"عدد الأجهزة المركبة: {warehouse.count_installed}",
            f"عدد الأجهزة المعطلة: {warehouse.count_damaged}",
        ]
        
        # إعادة تشكيل كل عنصر بالعربية وإضافته في صف واحد
        reshaped_stats = [
            Paragraph(get_display(arabic_reshaper.reshape(s)), ParagraphStyle("stat_text", fontName="Janna", fontSize=11, alignment=2))
            for s in stats
        ]
        
        # تقسيمها إلى صفوف حسب عدد الأعمدة (مثلاً 3 أعمدة)
        row_size = 2
        stat_rows = [reshaped_stats[i:i+row_size] for i in range(0, len(reshaped_stats), row_size)]
        
        # إنشاء الجدول
        stats_table = Table(stat_rows, colWidths=[380] * row_size, hAlign='RIGHT')
        stats_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, -1), "Janna"),
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        
        content.append(stats_table)
        content.append(Spacer(1, 20))

        
        headers = [
            "ملاحظات",
            "المنطقة",
            "تاريخ التركيب",
            "تاريخ النقل",
            "المسؤول",
            "IP",
            "الحالة",
            "الفئة",
            "الاسم",
            "الرقم التسلسلي"
        ]
        data = [[get_display(arabic_reshaper.reshape(h)) for h in headers]]

        for device in queryset:
            row = [
                device.notes or "",
                device.zone.name if device.zone else "المخزن",
                device.responsible_person or "",
                device.installation_date.strftime("%Y-%m-%d") if device.installation_date else "",
                device.transfer_date.strftime("%Y-%m-%d") if device.transfer_date else "",
                device.ip_address or "",
                dict(Device.DEVICE_STATUS_CHOICES).get(device.status, ""),
                device.device_category.name if device.device_category else "",
                device.name,
                device.serial_number
            ]
            row = [get_display(arabic_reshaper.reshape(str(col))) if any('\u0600' <= c <= '\u06FF' for c in str(col)) else str(col) for col in row]
            data.append(row)

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, -1), "Janna"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ]))
        content.append(table)

        doc.build(content)
        filename = f"devices_{slugify(warehouse.name)}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return HttpResponse(buffer.getvalue(), content_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        warehouse = self.object
        devices = self.get_filtered_devices().select_related('device_category', 'zone')
        context.update({
            'devices': devices,
            'filter_status': self.request.GET.get('status'),
            'filter_zone': self.request.GET.get('zone'),
            'search_query': self.request.GET.get('q'),
            'zones': Zone.objects.filter(contract=warehouse.contract),
            'total_devices': devices.count(),
            'category_counts': devices.values('device_category__name').annotate(total=Count('serial_number'))
        })
        return context

    def get_filtered_devices(self):
        warehouse = self.object
        devices = warehouse.devices.all()
        status = self.request.GET.get('status')
        zone_id = self.request.GET.get('zone')
        query = self.request.GET.get('q')

        if status:
            devices = devices.filter(status=status)
        if zone_id == 'warehouse':
            devices = devices.filter(current_location='warehouse')
        elif zone_id:
            devices = devices.filter(zone_id=zone_id)
        if query:
            devices = devices.filter(name__icontains=query)

        return devices

    def export_to_excel(self, queryset):
        warehouse = self.object
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Devices"

        ws.merge_cells("A1:E1")
        ws["A1"] = f"Warehouse: {warehouse.name} ({warehouse.location})"
        ws["A1"].font = Font(bold=True)

        if warehouse.contract:
            ws.merge_cells("A2:E2")
            ws["A2"] = f"Linked Contract: {warehouse.contract.name} ({warehouse.contract.contract_number})"
            ws["A2"].font = Font(bold=True)

        ws.append([])

        stats = [
            ("Total Zones", warehouse.count_zones),
            ("Total Devices", warehouse.count_devices),
            ("In Warehouse", warehouse.count_in_warehouse),
            ("Installed", warehouse.count_installed),
            ("Damaged", warehouse.count_damaged),
        ]
        ws.append(["Statistics", "Count"])
        for label, count in stats:
            ws.append([label, count])

        ws.append([])

        headers = ["Serial", "Name", "Category", "Status", "IP", "Transfer Date", "Installation Date", "Responsible", "Notes"]
        ws.append(headers)

        for device in queryset:
            ws.append([
                device.serial_number,
                device.name,
                device.device_category.name if device.device_category else "",
                dict(Device.DEVICE_STATUS_CHOICES).get(device.status, ""),
                device.ip_address,
                device.transfer_date.strftime("%Y-%m-%d") if device.transfer_date else "",
                device.installation_date.strftime("%Y-%m-%d") if device.installation_date else "",
                device.responsible_person,
                device.notes or ""
            ])

        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"devices_{warehouse.name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    def export_to_pdf(self, queryset):
        warehouse = self.object
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)

        font_path = os.path.join(settings.STATIC_ROOT, "contracts/fonts/Janna LT Bold/Janna LT Bold.ttf")
        pdfmetrics.registerFont(TTFont("Janna", font_path))

        content = []
        svg_path = os.path.join(settings.STATIC_ROOT, "contracts/img/logo-1.svg")
        drawing = svg2rlg(svg_path)
        drawing.scale(1, 1)
        drawing.hAlign = "LEFT"

        title_text = get_display(arabic_reshaper.reshape(f"تقرير أجهزة العقد  : {warehouse.contract.name} - {warehouse.contract.contract_number}"))
        title_paragraph = Paragraph(title_text, ParagraphStyle(name="Title", fontName="Janna", fontSize=20, alignment=1))

        title_table = Table([[drawing, title_paragraph]], colWidths=[80, 650], hAlign='RIGHT')
        title_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, 0), "RIGHT"),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ]))

        content.append(title_table)
        content.append(Spacer(1, 25))

        stats = [
            f"عدد الأجهزة في المخزن: {warehouse.count_in_warehouse}",
            f"المخزن: {warehouse.name} - {warehouse.location}",
            f"إجمالي المناطق: {warehouse.count_zones}",
            f"إجمالي الأجهزة: {warehouse.count_devices}",
            f"عدد الأجهزة المعطلة: {warehouse.count_damaged}",
            f"عدد الأجهزة المركبة: {warehouse.count_installed}"
        ]

        reshaped_stats = [
            Paragraph(get_display(arabic_reshaper.reshape(s)), ParagraphStyle("stat_text", fontName="Janna", fontSize=11, alignment=2))
            for s in stats
        ]

        row_size = 2
        stat_rows = [reshaped_stats[i:i + row_size] for i in range(0, len(reshaped_stats), row_size)]
        stats_table = Table(stat_rows, colWidths=[380] * row_size, hAlign='RIGHT')
        stats_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, -1), "Janna"),
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))

        content.append(stats_table)
        content.append(Spacer(1, 20))

        headers = ["ملاحظات", "المنطقة", "تاريخ التركيب", "تاريخ النقل", "المسؤول", "IP", "الحالة", "الفئة", "الاسم", "الرقم التسلسلي"]
        data = [[get_display(arabic_reshaper.reshape(h)) for h in headers]]

        for device in queryset:
            row = [
                device.notes or "",
                device.zone.name if device.zone else "المخزن",
                device.installation_date.strftime("%Y-%m-%d") if device.installation_date else "",
                device.transfer_date.strftime("%Y-%m-%d") if device.transfer_date else "",
                device.responsible_person or "",
                device.ip_address or "",
                dict(Device.DEVICE_STATUS_CHOICES).get(device.status, ""),
                device.device_category.name if device.device_category else "",
                device.name,
                device.serial_number
            ]
            row = [get_display(arabic_reshaper.reshape(str(col))) if any("\u0600" <= c <= "\u06FF" for c in str(col)) else str(col) for col in row]
            data.append(row)

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, -1), "Janna"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT")
        ]))

        content.append(table)
        doc.build(content)

        filename = f"devices_{warehouse.name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return HttpResponse(buffer.getvalue(), content_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

class DeviceImportView(AuthViewMixin, View):
    template_name = "contracts/devices/import_devices.html"

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        excel_file = request.FILES.get("file")
        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect("import_devices")

        try:
            df = pd.read_excel(excel_file, engine="openpyxl")

            for index, row in df.iterrows():
                category_name = str(row.get("device_category") or "").strip()
                warehouse_name = str(row.get("warehouse_name") or "").strip()
                zone_name = str(row.get("zone_name") or "").strip()

                if not category_name or not warehouse_name:
                    continue

                category, _ = DeviceCategory.objects.get_or_create(name=category_name)

                contract_name = f"AutoContract-{warehouse_name}"
                contract, _ = Contract.objects.get_or_create(
                    name=contract_name,
                    defaults={
                        "contract_number": f"CN-{warehouse_name[:5]}-{index}",
                        "start_date": timezone.now().date(),
                        "notes": "Auto-generated contract for import"
                    }
                )

                warehouse, _ = Warehouse.objects.get_or_create(
                    name=warehouse_name,
                    defaults={"location": "Auto-Imported", "contract": contract}
                )

                zone = None
                if zone_name:
                    zone, _ = Zone.objects.get_or_create(
                        name=zone_name,
                        contract=contract,
                        defaults={"notes": "Auto-created via import"}
                    )

                ContractItem.objects.get_or_create(
                    contract=contract,
                    category=category,
                    defaults={"quantity": 1, "notes": "Auto-added on import"}
                )

                Device.objects.update_or_create(
                    serial_number=str(row.get("serial_number")).strip(),
                    defaults={
                        "name": str(row.get("name") or "").strip(),
                        "invoice_number": str(row.get("invoice_number") or "").strip(),
                        "device_category": category,
                        "warehouse": warehouse,
                        "zone": zone,
                        "status": str(row.get("status") or "").strip(),
                        "current_location": str(row.get("current_location") or "").strip(),
                        "ip_address": row.get("ip_address") or None,
                        "responsible_person": str(row.get("responsible_person") or "").strip(),
                        "transfer_date": row.get("transfer_date") if not pd.isna(row.get("transfer_date")) else None,
                        "installation_date": row.get("installation_date") if not pd.isna(row.get("installation_date")) else None,
                        "notes": str(row.get("notes") or "").strip(),
                    }
                )

            messages.success(request, "Devices imported successfully.")

        except Exception as e:
            messages.error(request, f"An error occurred during import: {str(e)}")

        return redirect("import_devices")

class DeviceFormView(AuthViewMixin, FormView):
    model = Device
    form_class = DeviceForm
    template_name = 'contracts/devices/form.html'

    def dispatch(self, request, *args, **kwargs):
        self.device = None 
        if self.kwargs.get('pk'):
            self.device = get_object_or_404(Device, pk=self.kwargs['pk'])
            self.warehouse = self.device.warehouse
        else:
            self.warehouse = get_object_or_404(Warehouse, pk=self.kwargs.get('warehouse_id'))

        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['contract'] = self.warehouse.contract
        if self.device:
            kwargs['instance'] = self.device
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.method == 'POST':
            context['formset'] = DevicePropertyFormSet(self.request.POST, instance=self.device, prefix='properties')
        else:
            context['formset'] = DevicePropertyFormSet(instance=self.device, prefix='properties')
        context['warehouse_id'] = self.warehouse.pk
        context['device'] = self.device
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        form.instance.warehouse = self.warehouse

        if form.is_valid() and formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()

            action = 'updated' if self.device else 'added'
            messages.success(self.request, f"Device '{form.instance.name}' has been successfully {action}.")
            return redirect(self.get_success_url())

        context['form'] = form
        return self.render_to_response(context)

    def get_success_url(self):
        return reverse_lazy('warehouse_detail', kwargs={'pk': self.warehouse.pk})

class DeviceDeleteView(AuthViewMixin, DeleteView):
    model = Device
    template_name = 'contracts/devices/delete.html'
    context_object_name = 'device'
    pk_url_kwarg = 'pk'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['warehouse_id'] = self.object.warehouse.pk
        return context

    def get_success_url(self):
        return reverse_lazy('warehouse_detail', kwargs={'pk': self.object.warehouse.pk})

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        device_name = self.object.name
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f"Device '{device_name}' has been deleted successfully.")
        return response

class DeviceDetailView(AuthViewMixin, DetailView):
    model = Device
    template_name = 'contracts/devices/detail.html'
    context_object_name = 'device'

@require_POST
@csrf_exempt
def update_device_status(request, pk):
    device = get_object_or_404(Device, pk=pk)
    new_status = request.POST.get('status')

    if new_status == 'installed':
        zone_id = request.POST.get('zone')
        if not zone_id:
            messages.error(request, "You must select a zone before installation.")
            return redirect('device_detail', pk=device.pk)

        try:
            zone = Zone.objects.get(pk=zone_id)
            device.zone = zone
        except Zone.DoesNotExist:
            messages.error(request, "The selected zone does not exist.")
            return redirect('device_detail', pk=device.pk)

    elif new_status == 'available':
        device.zone = None

    if new_status in dict(Device.DEVICE_STATUS_CHOICES):
        device.status = new_status
        device.save()
        messages.success(request, f"Device status updated to: {dict(Device.DEVICE_STATUS_CHOICES).get(new_status)}")

        if new_status == 'damaged':
            return redirect('maintenance_add', device_id=device.pk)

        return redirect('warehouse_detail', pk=device.warehouse.pk)

    messages.error(request, "Invalid status selected.")
    return redirect('device_detail', pk=device.pk)

# Maintenance Views
class MaintenanceListView(AuthViewMixin, ExportMixin, ListView):
    model = MaintenanceCard
    template_name = 'contracts/maintenance/list.html'
    context_object_name = 'maintenance_cards'

    def get_queryset(self):
        queryset = super().get_queryset().select_related('device', 'device__zone', 'device__zone__contract')

        # Get filters
        contract_number = self.request.GET.get('contract')
        zone_id = self.request.GET.get('zone')
        category_id = self.request.GET.get('category')
        status = self.request.GET.get('status')

        # Get contract
        if contract_number:
            contract = Contract.objects.filter(contract_number=contract_number).first()
        else:
            contract = Contract.objects.last()

        if contract:
            queryset = queryset.filter(device__zone__contract=contract)

        # Zone filter
        if zone_id == "warehouse":
            queryset = queryset.filter(device__current_location="warehouse")
        elif zone_id:
            queryset = queryset.filter(device__zone_id=zone_id)

        # Category filter
        if category_id:
            queryset = queryset.filter(device__device_category_id=category_id)

        # Status filter
        if status and status.lower() != "none":
            queryset = queryset.filter(device__status=status)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        contract_number = self.request.GET.get('contract')
        zone_id = self.request.GET.get('zone')
        category_id = self.request.GET.get('category')
        status = self.request.GET.get('status')

        # Get current contract (to list zones)
        if contract_number:
            contract = Contract.objects.filter(contract_number=contract_number).first()
        else:
            contract = Contract.objects.last()

        context.update({
            'contracts': Contract.objects.all(),
            'zones': Zone.objects.filter(contract=contract) if contract else Zone.objects.none(),
            'categories': DeviceCategory.objects.all(),

            'filter_contract': contract_number or (contract.contract_number if contract else None),
            'filter_zone': zone_id,
            'filter_category': category_id,
            'filter_status': status,
        })

        return context

class MaintenanceCreateView(AuthViewMixin, CreateView):
    model = MaintenanceCard
    form_class = MaintenanceCardForm
    template_name = 'contracts/maintenance/form.html'

    def dispatch(self, request, *args, **kwargs):
        self.device = get_object_or_404(Device, pk=self.kwargs['device_id'])
        return super().dispatch(request, *args, **kwargs)

    def get_initial(self):
        return {'device': self.device}

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['device'] = self.device
        return context

    def form_valid(self, form):
        form.instance.device = self.device
        messages.success(self.request, "Maintenance card created successfully.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "There was an error creating the maintenance card.")
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('device_detail', kwargs={'pk': self.device.pk})

class MaintenanceUpdateView(AuthViewMixin, UpdateView):
    model = MaintenanceCard
    form_class = MaintenanceCardForm
    template_name = 'contracts/maintenance/form.html'
    pk_url_kwarg = 'pk'

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.device = self.object.device
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['device'] = self.device
        return context

    def form_valid(self, form):
        messages.success(self.request, "Maintenance card updated successfully.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Error updating the maintenance card.")
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('device_detail', kwargs={'pk': self.device.pk})

class MaintenanceDeleteView(AuthViewMixin, DeleteView):
    model = MaintenanceCard
    template_name = 'contracts/maintenance/delete.html'
    pk_url_kwarg = 'pk'

    def get_success_url(self):
        return reverse_lazy('maintenance_list')


# Coordination Views
class CoordinationListView(AuthViewMixin, ExportMixin, ListView):
    model = CoordinationRequest
    template_name = 'contracts/coordination/list.html'
    context_object_name = 'coordination'
    paginate_by = 25

    def get_queryset(self):
        contract_id = self.request.GET.get('contract')
        zone_id = self.request.GET.get('zone')
        queryset = CoordinationRequest.objects.all()

        if contract_id:
            queryset = queryset.filter(zone__contract_id=contract_id)
        else:
            last_contract = Contract.objects.last()
            if last_contract:
                queryset = queryset.filter(zone__contract=last_contract)

        if zone_id and zone_id.isdigit():
            queryset = queryset.filter(zone_id=int(zone_id))

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        last_contract = Contract.objects.last()
        contract_id = self.request.GET.get('contract') or (last_contract.contract_number if last_contract else None)
        zone_id = self.request.GET.get('zone')

        context.update({
            'contracts': Contract.objects.all(),
            'zones': Zone.objects.filter(contract_id=contract_id) if contract_id else Zone.objects.none(),
            'filter_contract': contract_id,
            'filter_zone': zone_id if zone_id and zone_id.isdigit() else ''
        })

        return context

class CoordinationCreateView(AuthViewMixin, CreateView):
    model = CoordinationRequest
    form_class = CoordinationRequestForm
    template_name = 'contracts/coordination/form.html'
    success_url = reverse_lazy('coordination_list')

    def form_valid(self, form):
        messages.success(self.request, "Coordination request created successfully.")
        return super().form_valid(form)

class CoordinationUpdateView(AuthViewMixin, UpdateView):
    model = CoordinationRequest
    form_class = CoordinationRequestForm
    template_name = 'contracts/coordination/form.html'
    success_url = reverse_lazy('coordination_list')

    def form_valid(self, form):
        messages.success(self.request, "Coordination request updated successfully.")
        return super().form_valid(form)

class CoordinationDeleteView(AuthViewMixin, DeleteView):
    model = CoordinationRequest
    template_name = 'contracts/coordination/delete.html'
    success_url = reverse_lazy('coordination_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Coordination request deleted successfully.")
        return super().delete(request, *args, **kwargs)

class TaskListView(AuthViewMixin, ListView):
    model = Task
    template_name = 'contracts/tasks/list.html'
    context_object_name = 'tasks'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().select_related('zone', 'zone__contract')

        contract_id = self.request.GET.get('contract')
        zone_id = self.request.GET.get('zone')
        status = self.request.GET.get('status')

        if contract_id:
            queryset = queryset.filter(zone__contract_id=contract_id)
        else:
            last_contract = Contract.objects.last()
            if last_contract:
                queryset = queryset.filter(zone__contract=last_contract)

        if zone_id:
            queryset = queryset.filter(zone_id=zone_id)

        if status:
            queryset = queryset.filter(status=status)

        return queryset.order_by('deadline')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contract_id = self.request.GET.get('contract')

        context['contracts'] = Contract.objects.all()

        # تحديد العقد الحالي أو الأخير
        if contract_id:
            context['filter_contract'] = str(contract_id)
            context['zones'] = Zone.objects.filter(contract_id=contract_id)
        else:
            last_contract = Contract.objects.last()
            context['filter_contract'] = last_contract.pk if last_contract else None
            context['zones'] = Zone.objects.filter(contract=last_contract) if last_contract else []

        context['selected_zone'] = self.request.GET.get('zone')
        context['selected_status'] = self.request.GET.get('status')

        return context

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        contract_id = self.request.GET.get('contract')
        zone_id = self.request.GET.get('zone')
        status = self.request.GET.get('status')

        last_contract = Contract.objects.last()

        context.update({
            'contracts': Contract.objects.all(),
            'zones': Zone.objects.filter(contract_id=contract_id) if contract_id else Zone.objects.filter(contract=last_contract) if last_contract else Zone.objects.none(),
            'filter_contract': contract_id or (last_contract.pk if last_contract else None),
            'selected_zone': zone_id,
            'selected_status': status,
        })

        return context


@require_GET
@csrf_exempt
def task_change_status(request, pk, status):
    task = get_object_or_404(Task, pk=pk)
    valid_statuses = [s[0] for s in Task.TASK_STATUS_CHOICES]

    if status in valid_statuses:
        task.status = status
        if status == 'completed':
            task.actual_delivery_date = timezone.now().date()
        else:
            task.actual_delivery_date = None
        task.save()
        messages.success(request, f"Task status updated to {task.get_status_display()}.")
    else:
        messages.error(request, "Invalid status selected.")

    # ⏎ إعادة توجيه مع نفس الفلاتر
    base_url = reverse('task_list')
    query_string = request.META.get('QUERY_STRING', '')
    return redirect(f"{base_url}?{query_string}")

class TaskCreateView(AuthViewMixin, CreateView):
    model = Task
    form_class = TaskForm
    template_name = 'contracts/tasks/form.html'

    def form_valid(self, form):
        messages.success(self.request, f"Task '{form.instance.name}' has been created successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('task_list')


class TaskUpdateView(AuthViewMixin, UpdateView):
    model = Task
    form_class = TaskForm
    template_name = 'contracts/tasks/form.html'

    def form_valid(self, form):
        messages.success(self.request, f"Task '{form.instance.name}' has been updated successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('task_list')


class TaskDeleteView(AuthViewMixin, DeleteView):
    model = Task
    template_name = 'contracts/tasks/delete.html'
    success_url = reverse_lazy('task_list')

    def delete(self, request, *args, **kwargs):
        task = self.get_object()
        messages.success(request, f"Task '{task.name}' has been deleted successfully.")
        return super().delete(request, *args, **kwargs)
